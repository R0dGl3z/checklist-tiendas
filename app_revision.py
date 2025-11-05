import streamlit as st
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import os

st.set_page_config(page_title="Revisión de Tienda", layout="centered")

st.title("📋 Revisión de Tienda - Danny Yo")

# ====== ENCABEZADO ======
st.subheader("Información general")
sucursal = st.text_input("Sucursal", "Perla 1")
fecha = st.date_input("Fecha", datetime.now())
hora_entrada = st.time_input("Hora de entrada")
hora_salida = st.time_input("Hora de salida")
empleados = st.text_area("Empleados en revisión (uno por línea)")

st.divider()

# ====== PREGUNTAS ======
st.subheader("Checklist de revisión")

preguntas = [
    "Verificar apertura de tienda de acuerdo a su horario",
    "Garantizar la limpieza permanente del local en general",
    "Comprobar que los mostradores y vitrinas estén limpias",
]

respuestas = []
comentarios = []
evidencias = []

for i, pregunta in enumerate(preguntas, start=1):
    st.markdown(f"### {i}. {pregunta}")
    opcion = st.radio(
        f"Selecciona el resultado para la pregunta {i}:",
        ["Cumple", "No cumple", "N/A"],
        key=f"radio_{i}"
    )
    comentario = st.text_area(f"Comentarios (pregunta {i})", key=f"comentario_{i}")
    evidencia = st.file_uploader(f"Adjunta evidencia (imagen o PDF)", key=f"evidencia_{i}")

    respuestas.append(opcion)
    comentarios.append(comentario)
    evidencias.append(evidencia)
    st.divider()

# ====== BOTÓN FINAL ======
if st.button("✅ Generar Excel"):
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Revisión"

    # Encabezado
    ws["A1"] = "Sucursal"; ws["B1"] = sucursal
    ws["A2"] = "Fecha"; ws["B2"] = fecha.strftime("%d/%m/%Y")
    ws["A3"] = "Hora entrada"; ws["B3"] = hora_entrada.strftime("%I:%M %p")
    ws["A4"] = "Hora salida"; ws["B4"] = hora_salida.strftime("%I:%M %p")
    ws["A5"] = "Empleados revisados"; ws["B5"] = empleados

    ws.append([""])
    ws.append(["Pregunta", "Resultado", "Comentario", "Evidencia"])

    # Guardar respuestas
    for i, pregunta in enumerate(preguntas):
        resultado = respuestas[i] if i < len(respuestas) else ""
        comentario = comentarios[i] if i < len(comentarios) else ""
        evidencia = evidencias[i] if i < len(evidencias) else None

        ws.append([pregunta, resultado, comentario])

        # Insertar imagen si hay evidencia
        if evidencia is not None:
            file_path = f"temp_evidencia_{i}.png"
            with open(file_path, "wb") as f:
                f.write(evidencia.getbuffer())
            try:
                img = Image(file_path)
                img.width = 200
                img.height = 120
                ws.add_image(img, f"D{ws.max_row}")
            except Exception as e:
                st.warning(f"No se pudo insertar la imagen: {e}")

    # Guardar archivo
    file_name = f"Revision_{sucursal}_{fecha.strftime('%Y%m%d')}.xlsx"
    wb.save(file_name)

    with open(file_name, "rb") as f:
        st.download_button(
            label="📥 Descargar reporte Excel",
            data=f,
            file_name=file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    st.success("✅ Excel generado correctamente.")