import streamlit as st
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from datetime import datetime
import os
import tempfile

# ========================
# CONFIGURACIÓN INICIAL
# ========================
st.set_page_config(page_title="Check List Tiendas", layout="wide")
PLANTILLA = "CHECK LIST Tiendas.xlsx"

# ========================
# CONFIGURACIÓN PARA CELULAR
# ========================
st.markdown("""
<style>
/* Centrar contenido y mejorar espaciado en móvil */
.block-container {
    padding-top: 1rem;
    padding-bottom: 2rem;
    padding-left: 1rem;
    padding-right: 1rem;
}
/* Hacer botones más grandes */
button[kind="primary"] {
    font-size: 18px !important;
    padding: 10px 20px !important;
}
/* Mejorar legibilidad del texto */
textarea, input, label {
    font-size: 16px !important;
}
</style>
""", unsafe_allow_html=True)

# ========================
# FORMULARIO DE ENCABEZADO
# ========================
st.title("📋 Check List Tiendas - Danny Yo")

sucursal = st.text_input("Sucursal", "")
fecha = st.date_input("Fecha", datetime.now())
hora_entrada = st.text_input("Hora de entrada (ej. 2:13 p.m.)")
hora_salida = st.text_input("Hora de salida (ej. 4:05 p.m.)")
empleados = st.text_area("Empleados en revisión (uno por línea)")

st.divider()

# ========================
# LEER PREGUNTAS (F17:F178 ignorando encabezados)
# ========================
wb = load_workbook(PLANTILLA)
ws = wb.active

preguntas = []
for row in range(17, 179):
    valor = ws[f"F{row}"].value
    if valor and isinstance(valor, str):
        texto = valor.strip()

        # Detectar encabezados de bloque (PERSONAL, CALIDAD, etc.)
        mayus = texto.upper() == texto
        pocas_palabras = len(texto.split()) <= 4
        no_tiene_puntuacion = not any(c in texto for c in [".", "?", ":", ";"])

        # Si cumple las tres condiciones, se considera encabezado
        if mayus and pocas_palabras and no_tiene_puntuacion:
            continue

        preguntas.append((row, texto))

if not preguntas:
    st.error("⚠️ No se detectaron preguntas válidas entre F17 y F178. Revisa la plantilla.")
    st.stop()

# ========================
# FORMULARIO DE PREGUNTAS
# ========================
respuestas = []
comentarios = []
evidencias = []

st.subheader("Responde el checklist:")

for i, (fila, pregunta) in enumerate(preguntas, start=1):
    st.markdown(f"### {i}. {pregunta}")
    opcion = st.radio(
        "Resultado:",
        ["Cumple", "No cumple", "N/A"],
        key=f"resp_{i}",
        horizontal=True,
        index=None  # sin preselección
    )
    comentario = st.text_area("Comentario:", key=f"coment_{i}")
    evidencia = st.file_uploader("Adjuntar evidencia (imagen opcional)", key=f"evid_{i}")

    respuestas.append(opcion)
    comentarios.append(comentario)
    evidencias.append(evidencia)
    st.divider()

# ========================
# CAMPOS FINALES
# ========================
st.subheader("Sección final del checklist")

observaciones = st.text_area("📝 Observaciones:")
anexo_fotos = st.file_uploader("📸 Anexo fotos de más áreas (puedes subir varias)", accept_multiple_files=True)
persona_visita = st.text_input("👤 Persona que realizó la visita:")
seguimiento = st.text_input("📨 Se informa para su seguimiento a:")

# ========================
# GENERAR EXCEL (sin validación obligatoria)
# ========================
if st.button("✅ Generar Check List"):
    wb = load_workbook(PLANTILLA)
    ws = wb.active

    # ====== ENCABEZADO ======
    ws["H2"] = sucursal
    ws["H5"] = fecha.strftime("%d/%m/%Y")
    ws["J6"] = hora_entrada
    ws["J8"] = hora_salida
    ws["F12"] = empleados

    temp_dir = tempfile.mkdtemp()
    temp_files = []

    # ====== PREGUNTAS ======
    for idx, (fila, pregunta) in enumerate(preguntas):
        resp = respuestas[idx]
        coment = comentarios[idx]
        evid = evidencias[idx]

        # Limpiar celdas previas
        ws[f"B{fila}"].value = ""
        ws[f"C{fila}"].value = ""
        ws[f"D{fila}"].value = ""
        ws[f"E{fila}"].value = ""

        # Solo escribir si respondió
        if resp == "Cumple":
            ws[f"B{fila}"].value = "X"
        elif resp == "No cumple":
            ws[f"C{fila}"].value = "X"
        elif resp == "N/A":
            ws[f"D{fila}"].value = "X"

        # Comentario (si existe)
        if coment:
            ws[f"E{fila}"].value = coment

        # Evidencia (si existe)
        if evid is not None:
            file_path = os.path.join(temp_dir, f"evidencia_{idx}.png")
            with open(file_path, "wb") as f:
                f.write(evid.getbuffer())
            temp_files.append(file_path)
            try:
                img = Image(file_path)
                img.width = 150
                img.height = 120
                ws.add_image(img, f"G{fila}")
            except Exception as e:
                st.warning(f"No se pudo insertar imagen en fila {fila}: {e}")

    # ====== CAMPOS FINALES ======
    ws["F183"] = observaciones

    # Fotos de anexo (rango F187:F192)
    start_row = 187
    for i, foto in enumerate(anexo_fotos):
        if i >= 6:
            break
        file_path = os.path.join(temp_dir, f"anexo_{i}.png")
        with open(file_path, "wb") as f:
            f.write(foto.getbuffer())
        temp_files.append(file_path)
        try:
            img = Image(file_path)
            img.width = 220
            img.height = 150
            ws.add_image(img, f"F{start_row + i}")
        except Exception as e:
            st.warning(f"No se pudo insertar imagen anexa {i+1}: {e}")

    # Persona que realizó la visita (mantiene texto)
    original_persona = ws["F193"].value or "PERSONA QUE REALIZO LA VISITA:"
    ws["F193"] = f"{original_persona.split(':')[0]}: {persona_visita}"

    # Se informa para su seguimiento a (mantiene texto)
    original_seguimiento = ws["F195"].value or "SE INFORMA PARA SU SEGUIMIENTO A:"
    ws["F195"] = f"{original_seguimiento.split(':')[0]}: {seguimiento}"

    # ====== GUARDAR ARCHIVO ======
    nombre_salida = f"CHECKLIST_{sucursal}_{fecha.strftime('%Y%m%d')}.xlsx"
    wb.save(nombre_salida)

    # Limpiar temporales
    for fpath in temp_files:
        if os.path.exists(fpath):
            os.remove(fpath)
    os.rmdir(temp_dir)

    with open(nombre_salida, "rb") as f:
        st.download_button(
            label="📥 Descargar Check List generado",
            data=f,
            file_name=nombre_salida,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    st.success("✅ ¡Check List generado correctamente!")
